import os
import re
import openpyxl
from collections import OrderedDict


def _parse_num(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except:
        return None


def parse_sheet(ws):
    """Parse a single sheet, detecting its type automatically."""
    info = OrderedDict()
    info["name"] = ws.title

    # Detect if it's a variable sheet (V01-V15)
    var_name = None
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                m = re.match(r"(V\d{2}|I\d{2})\s*[-–—]\s*(.+)", cell.strip())
                if m:
                    var_name = cell.strip()
                    break
        if var_name:
            break
    info["var_name"] = var_name

    # Detect data type (Soma/Amplitude format)
    has_control_chart = False
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row and row[0] and isinstance(row[0], str) and "Período" in row[0]:
            has_control_chart = True
            break

    if not has_control_chart:
        info["type"] = "other"
        info["stats"] = {}
        info["headers"] = []
        info["data"] = []
        return info

    info["type"] = "control_chart"

    # Extract stats
    stats = {}
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        cells = [c for c in row if c is not None]
        if not cells:
            continue
        label = str(cells[0]).strip() if cells else ""
        if "Média" in label and len(cells) >= 3:
            stats["media_soma"] = _parse_num(cells[1])
            stats["media_amp"] = _parse_num(cells[2])
        elif "Desvio" in label and len(cells) >= 3:
            stats["dp_soma"] = _parse_num(cells[1])
            stats["dp_amp"] = _parse_num(cells[2])
        elif "Máximo" in label and len(cells) >= 3:
            stats["max_soma"] = _parse_num(cells[1])
            stats["max_amp"] = _parse_num(cells[2])
        elif "Mínimo" in label and len(cells) >= 2:
            vals = [_parse_num(c) for c in cells[1:]]
            for v in vals:
                if v is not None:
                    stats["min_soma"] = vals[0] if len(vals) > 0 else None
                    stats["min_amp"] = vals[1] if len(vals) > 1 else None
                    break
    info["stats"] = stats

    # Find data start
    data_start = None
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        if row and row[0] and isinstance(row[0], str) and "Período" in row[0].strip():
            data_start = idx
            break

    if not data_start:
        info["headers"] = []
        info["data"] = []
        return info

    # Parse headers
    headers = []
    for cell in ws[data_start]:
        headers.append(str(cell.value).strip() if cell.value else "")
    info["headers"] = headers

    # Parse data rows
    data = []
    for row in ws.iter_rows(min_row=data_start + 1, values_only=True):
        if row[0] is None:
            continue
        periodo = str(row[0]).strip()
        if not re.match(r"^\d{4}[\.\-]\d{1,2}", periodo):
            continue

        entry = OrderedDict()
        entry["periodo"] = periodo
        for i, h in enumerate(headers[1:], start=1):
            val = row[i] if i < len(row) else None
            entry[h] = val
        data.append(entry)

    info["data"] = data
    info["total_rows"] = len(data)
    filled = sum(1 for d in data if d.get("Soma") is not None and str(d.get("Soma", "")).strip() != "")
    info["filled_rows"] = filled
    info["empty_rows"] = len(data) - filled

    if data:
        info["periodo_inicio"] = data[0]["periodo"]
        info["periodo_fim"] = data[-1]["periodo"]

    return info


def parse_all_sheets(filepath):
    """Parse all sheets from the Excel workbook."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    result = OrderedDict()
    result["_workbook"] = os.path.basename(filepath)
    result["_sheets"] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = parse_sheet(ws)
        result["_sheets"].append(sheet_info)

    return result


def analyze_indicators(parsed):
    """Compute statistics for each variable/indicator."""
    analysis = OrderedDict()

    for sheet in parsed["_sheets"]:
        label = sheet["var_name"] or sheet["name"]

        if sheet["type"] != "control_chart":
            continue

        info = OrderedDict()
        info["stats_header"] = sheet.get("stats", {})

        soma_vals = []
        amp_vals = []
        for d in sheet["data"]:
            try:
                s = d.get("Soma")
                if s is not None and str(s).strip() != "":
                    soma_vals.append(float(str(s).replace(",", ".")))
                a = d.get("Amplitude")
                if a is not None and str(a).strip() != "":
                    amp_vals.append(float(str(a).replace(",", ".")))
            except:
                pass

        info["total_linhas"] = sheet.get("total_rows", 0)
        info["preenchidas"] = sheet.get("filled_rows", 0)
        info["vazias"] = sheet.get("empty_rows", 0)
        info["periodo_inicio"] = sheet.get("periodo_inicio", "N/A")
        info["periodo_fim"] = sheet.get("periodo_fim", "N/A")

        if soma_vals:
            info["soma"] = {
                "min": min(soma_vals),
                "max": max(soma_vals),
                "media": round(sum(soma_vals) / len(soma_vals), 2),
            }
        if amp_vals:
            info["amplitude"] = {
                "min": round(min(amp_vals), 2),
                "max": round(max(amp_vals), 2),
                "media": round(sum(amp_vals) / len(amp_vals), 4),
            }

        analysis[label] = info

    return analysis


if __name__ == "__main__":
    import sys
    import json

    excel_path = "Risco de Integridade - Controle Estatistico PRGA.xlsx"
    if not os.path.exists(excel_path):
        print(f"Arquivo não encontrado: {excel_path}")
        sys.exit(1)

    parsed = parse_all_sheets(excel_path)
    analysis = analyze_indicators(parsed)

    print(f"Arquivo: {parsed['_workbook']}")
    print(f"Abas: {len(parsed['_sheets'])}")
    for s in parsed["_sheets"]:
        nome = s["var_name"] or s["name"]
        tipo = s["type"]
        dados = len(s.get("data", []))
        print(f"  [{tipo:>15s}] {nome}: {dados} linhas")

    print(f"\nIndicadores analisados: {len(analysis)}")
    for label, info in analysis.items():
        print(f"\n  {label}")
        print(f"    Período: {info['periodo_inicio']} a {info['periodo_fim']}")
        print(f"    Linhas: {info['preenchidas']} preenchidas / {info['vazias']} vazias")
        if "soma" in info:
            print(f"    Soma: min={info['soma']['min']}, max={info['soma']['max']}, média={info['soma']['media']}")
        if "amplitude" in info:
            print(f"    Amplitude: min={info['amplitude']['min']}, max={info['amplitude']['max']}")
