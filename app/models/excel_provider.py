import os, re, json, time
import requests
import browser_cookie3
import openpyxl
from datetime import datetime
from collections import OrderedDict

from app.models.default_data import (
    YEARS, dashboardData as fallbackData, quadStatus as fallbackQuadStatus,
    anualStatus as fallbackAnualStatus, integrityPairs, yearColors
)

SHAREPOINT_URL = (
    "https://caesbdfgovbr.sharepoint.com/:x:/r/sites/PRGC2/_layouts/15/Doc.aspx"
    "?sourcedoc=%7BD33B7D0A-8470-4A52-8C4E-03F30ACC8731%7D"
    "&file=Risco%20de%20Integridade%20-%20Controle%20Estatistico%20PRGA.xlsx"
    "&action=default&mobileredirect=true"
)
EXCEL_CACHE = os.path.join(os.path.dirname(__file__), "..", "..", ".excel_cache.xlsx")
DATA_CACHE = os.path.join(os.path.dirname(__file__), "..", "..", ".data_cache.json")
CACHE_TTL = 300

VAR_LABELS = {
    "V01": "Denúncias de assédio moral", "V02": "Total de denúncias",
    "V03": "Denúncias de assédio sexual", "V04": "Denúncias apuradas e tratadas",
    "V05": "Contratos enviados à CGDF", "V06": "Contratos enquadrados",
    "V07": "Empregados/membros treinados", "V08": "Total de empregados-membros",
    "V09": "Evasões", "V10": "Vagas ofertadas",
    "V11": "Denúncias de nepotismo", "V12": "Denúncias de conflito de interesses",
    "V13": "Denúncias procedentes", "V14": "Denúncias com penalidade",
    "V15": "Denúncias enviadas",
}

FORMULAS = {
    "I01": ("V01", "V02"), "I02": ("V03", "V02"), "I03": ("V04", "V02"),
    "I04": ("V06", "V05"), "I05": ("V07", "V08"), "I06": ("V09", "V10"),
    "I07": ("V11", "V02"), "I08": ("V12", "V02"), "I09": ("V13", "V02"),
    "I10": ("V14", "V15"),
}

QUAD_INDICATORS = {"I01", "I02", "I03", "I07", "I08"}
ANUAL_INDICATORS = {"I04", "I05", "I06", "I09", "I10"}

def month_to_quad(m):
    if 1 <= m <= 4: return 0
    if 5 <= m <= 8: return 1
    return 2

def _parse_num(val):
    if val is None: return None
    if isinstance(val, (int, float)): return val
    try: return float(str(val).strip().replace(",", "."))
    except: return None

def _extract_year_month(cell):
    if cell is None: return None, None
    if isinstance(cell, datetime): return str(cell.year), cell.month
    s = str(cell).strip()
    m = re.match(r"(\d{4})[\.\-](\d{1,2})(?:[\.\-]\d{2})?", s)
    if m: return m.group(1), int(m.group(2))
    return None, None

# ----- Download -----

def _chrome_cookies():
    merged = requests.cookies.RequestsCookieJar()
    for domain in ["caesbdfgovbr.sharepoint.com", "sharepoint.com", "microsoft.com", "microsoftonline.com"]:
        try:
            for c in browser_cookie3.chrome(domain_name=domain):
                merged.set(c.name, c.value, domain=c.domain, path=c.path)
        except: pass
    return merged

def _download_excel():
    try:
        cookies = _chrome_cookies()
        s = requests.Session()
        s.cookies.update(cookies)
        s.headers.update({"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) Chrome/130"})
        m_url = re.search(r'sourcedoc=%7B([a-f0-9\-]{36})%7D', SHAREPOINT_URL, re.I)
        uid = m_url.group(1).lower() if m_url else None
        if not uid:
            resp = s.get(SHAREPOINT_URL, allow_redirects=True, timeout=30)
            m = re.search(r'([a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12})', resp.text, re.I)
            uid = m.group(1).lower() if m else None
        if not uid: return None

        dl = f"https://caesbdfgovbr.sharepoint.com/sites/PRGC2/_layouts/15/download.aspx?UniqueId={uid}"

        # Follow SharePoint Forms-Based Authentication chain
        for _ in range(6):
            r = s.get(dl, allow_redirects=True, timeout=60)
            ct = r.headers.get("Content-Type", "")
            if "application/vnd" in ct or "application/zip" in ct or len(r.content) > 50000:
                with open(EXCEL_CACHE, "wb") as f:
                    f.write(r.content)
                return EXCEL_CACHE
            ear = re.search(r'name="ear_jwe" value="([^"]+)"', r.text)
            action = re.search(r'action="([^"]+)"', r.text)
            if ear and action:
                next_url = action.group(1).replace("&amp;", "&")
                r = s.post(next_url, data={"ear_jwe": ear.group(1)}, allow_redirects=True, timeout=30)
                ct = r.headers.get("Content-Type", "")
                if "application/vnd" in ct or "application/zip" in ct or len(r.content) > 50000:
                    with open(EXCEL_CACHE, "wb") as f:
                        f.write(r.content)
                    return EXCEL_CACHE
            else:
                break

        # Auth chain completed but file not delivered — try download URL once more
        r = s.get(dl, allow_redirects=True, timeout=60)
        ct = r.headers.get("Content-Type", "")
        if "application/vnd" in ct or "application/zip" in ct or len(r.content) > 50000:
            with open(EXCEL_CACHE, "wb") as f:
                f.write(r.content)
            return EXCEL_CACHE
    except:
        pass
    return None

# ----- Parse -----

def _parse_sheet(ws):
    data_start = None
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        if row and row[0] is not None and "Período" in str(row[0]).strip():
            data_start = idx; break
    if not data_start: return {}
    monthly = {}
    for row in ws.iter_rows(min_row=data_start + 1, values_only=True):
        if row[0] is None: continue
        year, month = _extract_year_month(row[0])
        if year is None: continue
        soma = _parse_num(row[1]) if len(row) > 1 else None
        if soma is None: continue
        if isinstance(row[1], str) and row[1].strip().lower() == "soma": continue
        monthly.setdefault(year, {})[month] = soma
    return monthly

def _load_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    var_data = {}
    for name in wb.sheetnames:
        m = re.match(r"(V\d{2})", name)
        if m:
            d = _parse_sheet(wb[name])
            if d: var_data[m.group(1)] = d
    wb.close()
    return var_data

# ----- Aggregate -----

def _agg_quad(monthly):
    r = {}
    for year, months in monthly.items():
        q = [0, 0, 0]
        for m, v in months.items():
            if v is not None: q[month_to_quad(m)] += v
        r[year] = q
    return r

def _agg_anual(monthly):
    return {year: sum(v for v in months.values() if v is not None) for year, months in monthly.items()}

# ----- Build response -----

def get_dashboard_data():
    if _cache_valid(DATA_CACHE):
        try:
            with open(DATA_CACHE) as f:
                cached = json.load(f)
                if cached: return cached
        except: pass

    excel_path = None
    if _cache_valid(EXCEL_CACHE): excel_path = EXCEL_CACHE
    else: excel_path = _download_excel()

    if excel_path and os.path.exists(excel_path):
        try:
            return _build_from_excel(excel_path)
        except Exception as e:
            print(f"[excel] Erro ao processar Excel: {e}")

    return None

def _cache_valid(path, ttl=CACHE_TTL):
    return os.path.exists(path) and (time.time() - os.path.getmtime(path)) < ttl

def _build_from_excel(path):
    var_data = _load_excel(path)
    if not var_data: return None

    quad, anual = {}, {}
    for vid, monthly in var_data.items():
        quad[vid] = _agg_quad(monthly)
        anual[vid] = _agg_anual(monthly)

    # Status from V02 reference
    ref_q = quad.get("V02", {})
    qs = {}
    for y in YEARS:
        vals = ref_q.get(y, [None]*3)
        qs[y] = ["completo" if v and v > 0 else ("parcial" if v is not None else "pendente") for v in vals]

    ref_a = anual.get("V02", {})
    aas = {}
    for y in YEARS:
        v = ref_a.get(y)
        aas[y] = "completo" if v else "pendente"

    # Count valid quadrimestres
    q_count = sum(1 for yd in ref_q.values() for v in yd if v and v > 0)

    # Build indicators
    dd = {}
    for code in sorted(FORMULAS):
        num, den = FORMULAS[code]
        meta = _indicator_meta(code)
        meta["coletasValidas"] = q_count if code in QUAD_INDICATORS else 0

        if code in QUAD_INDICATORS:
            meta["vars"] = {
                num: {"label": VAR_LABELS.get(num, num), "data": {y: quad.get(num, {}).get(y, [None]*3) for y in YEARS}},
                den: {"label": VAR_LABELS.get(den, den), "data": {y: quad.get(den, {}).get(y, [None]*3) for y in YEARS}},
            }
        else:
            meta["vars"] = {}
            for v in [num, den]:
                meta["vars"][v] = {
                    "label": VAR_LABELS.get(v, v),
                    "data": {y: anual.get(v, {}).get(y) for y in YEARS},
                }
            if code == "I09" and "V02" not in meta["vars"]:
                meta["vars"]["V02"] = {"label": "Total de denúncias", "data": {y: anual.get("V02", {}).get(y) for y in YEARS}}

        dd[code] = meta

    result = {
        "YEARS": YEARS, "dashboardData": dd,
        "quadStatus": qs, "anualStatus": aas,
        "integrityPairs": integrityPairs, "yearColors": yearColors,
    }

    try:
        with open(DATA_CACHE, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
    except: pass

    return result

def _indicator_meta(code):
    """Pega metadados do fallback e atualiza nome/fórmula."""
    fb = fallbackData.get(code, {})
    return {
        "name": fb.get("name", code),
        "formula": fb.get("formula", ""),
        "periodicidade": fb.get("periodicidade", ""),
        "granularidade": fb.get("granularidade", ""),
        "sentidoBom": fb.get("sentidoBom", ""),
        "coletasValidas": 0,
        "meta": fb.get("meta"),
        "limiteAceitavel": fb.get("limiteAceitavel"),
        "observacao": fb.get("observacao"),
    }
